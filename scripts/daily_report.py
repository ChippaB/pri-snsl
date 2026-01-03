"""
Daily Build Report Generator for Polytechnic Resources
Queries Supabase for previous day's scans, generates Excel report, emails to recipients.

Usage:
    python daily_report.py                  # Uses yesterday's date
    python daily_report.py 2025-12-17       # Specific date
    python daily_report.py --test           # Test mode (prints output, no email)
"""

import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import tempfile

# Third-party imports
try:
    from supabase import create_client, Client
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install supabase openpyxl")
    sys.exit(1)

# Optional: Google Sheets backup (gspread)
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False
    print("[WARN] gspread not installed - Google Sheets backup disabled")

# ============================================
# CONFIGURATION (from environment variables)
# ============================================

# Supabase
SUPABASE_URL = os.environ.get('SUPABASE_URL', 'https://ospedluufxgpfvqtznej.supabase.co')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9zcGVkbHV1ZnhncGZ2cXR6bmVqIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjU5ODgyMTUsImV4cCI6MjA4MTU2NDIxNX0.1AhtuANYs-eVrQIdW9gqt_KLhBxF4Vm0j6pqtrrJAag')

# Email Configuration
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_EMAIL = os.environ.get('SMTP_EMAIL', 'polytechnicresources.dev@gmail.com')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', 'jppg ytea oqzc lsdn')  # Gmail App Password

# Recipients (comma-separated in a single string)
REPORT_RECIPIENTS = os.environ.get('REPORT_RECIPIENTS', 'chip.brandner@gmail.com, ksilesky1@verizon.net')

# Report settings
PIECES_PER_BOX = int(os.environ.get('PIECES_PER_BOX', '100'))

# Google Sheets Backup Configuration
# GSHEET_CREDENTIALS_JSON: Base64-encoded service account JSON (from GitHub Secret)
# GSHEET_SPREADSHEET_ID: The ID from the Google Sheet URL
GSHEET_CREDENTIALS_JSON = os.environ.get('GSHEET_CREDENTIALS_JSON', '')
GSHEET_SPREADSHEET_ID = os.environ.get('GSHEET_SPREADSHEET_ID', '')
GSHEET_ENABLED = bool(GSHEET_CREDENTIALS_JSON and GSHEET_SPREADSHEET_ID and GSPREAD_AVAILABLE)


def get_supabase_client() -> Client:
    """Initialize Supabase client"""
    if not SUPABASE_KEY:
        raise ValueError("SUPABASE_KEY environment variable not set")
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def fetch_scans_for_date(supabase: Client, target_date: datetime) -> list:
    """
    Fetch all scans for a specific date (in EST/Eastern Time).
    Supabase stores timestamps in UTC, so we need to convert.
    """
    # Calculate UTC range for the target EST date
    # EST is UTC-5, so midnight EST = 5:00 AM UTC
    start_utc = target_date.replace(hour=5, minute=0, second=0, microsecond=0)
    end_utc = start_utc + timedelta(days=1)
    
    start_str = start_utc.strftime('%Y-%m-%dT%H:%M:%S+00:00')
    end_str = end_utc.strftime('%Y-%m-%dT%H:%M:%S+00:00')
    
    print(f"[INFO] Fetching scans from {start_str} to {end_str}")
    
    response = supabase.table('scans').select('*').gte(
        'created_at', start_str
    ).lt(
        'created_at', end_str
    ).order('operator_name').order('station_id').order('part_id').execute()
    
    return response.data if response.data else []


def format_serial_ranges(serials: list) -> str:
    """
    Convert a list of serial numbers into formatted ranges.
    Uses '-' for consecutive sequences, '/' for breaks.
    Example: ['MGCK173156', 'MGCK173157', 'MGCK173158', 'MGCK173170'] 
             -> 'MGCK173156-173158 / MGCK173170'
    """
    if not serials:
        return ""
    
    # Sort serials
    serials = sorted(set(serials))
    
    if len(serials) == 1:
        return serials[0]
    
    def extract_prefix_and_num(serial):
        """
        Extract prefix and trailing numeric portion.
        'PUL9000K29498' -> ('PUL9000K', 29498)
        'MGCK173156' -> ('MGCK', 173156)
        """
        import re
        # Find trailing digits
        match = re.match(r'^(.+?)(\d+)$', serial)
        if match:
            prefix = match.group(1)
            num = int(match.group(2))
            return prefix, num
        return serial, None
    
    # Build ranges
    ranges = []
    current_prefix = None
    current_start = None
    current_end = None
    
    for serial in serials:
        prefix, num = extract_prefix_and_num(serial)
        
        if num is None:
            # No numeric portion, output current range and add this as-is
            if current_start is not None:
                ranges.append(format_range(current_prefix, current_start, current_end))
                current_start = None
            ranges.append(serial)
            continue
        
        if current_start is None:
            # Start new range
            current_prefix = prefix
            current_start = num
            current_end = num
        elif prefix == current_prefix and num == current_end + 1:
            # Extend current range (consecutive)
            current_end = num
        else:
            # End current range, start new one
            ranges.append(format_range(current_prefix, current_start, current_end))
            current_prefix = prefix
            current_start = num
            current_end = num
    
    # Don't forget the last range
    if current_start is not None:
        ranges.append(format_range(current_prefix, current_start, current_end))
    
    return " / ".join(ranges)


def format_range(prefix: str, start: int, end: int) -> str:
    """Format a range like 'MGCK173156-173158' or 'MGCK173156' if single"""
    if start == end:
        return f"{prefix}{start}"
    else:
        return f"{prefix}{start}-{end}"

def extract_serial_header(serial: str) -> str:
    """
    Extract serial header prefix (everything before last 5 digits).
    'MGCK173156' -> 'MGCK1' (keeps 'MGCK1', drops '73156')
    'MGC2C12345' -> 'MGC2C'
    Handles edge cases like ranges stored as single values.
    """
    if not serial:
        return 'UNKNOWN'
    
    # If serial contains a hyphen (range stored as single value), use first part
    target = serial.split('-')[0]
    
    if len(target) <= 5:
        return target or 'UNKNOWN'
    
    # Take everything except last 5 characters
    return target[:-5]


def format_serial_numbers_only(serials: list) -> str:
    """
    Format serial numbers showing only the trailing numeric portions as ranges.
    'MGCK173156', 'MGCK173157', 'MGCK173158' -> '73156-73158'
    """
    if not serials:
        return ""
    
    # Extract just the last 5 digits from each serial
    numbers = []
    for serial in sorted(set(serials)):
        if serial and len(serial) > 5:
            try:
                num = int(serial[-5:])
                numbers.append(num)
            except ValueError:
                numbers.append(serial[-5:])  # Keep as string if not numeric
        else:
            numbers.append(serial)
    
    if not numbers:
        return ""
    
    # If all are integers, format as ranges
    if all(isinstance(n, int) for n in numbers):
        numbers = sorted(numbers)
        ranges = []
        start = numbers[0]
        end = numbers[0]
        
        for num in numbers[1:]:
            if num == end + 1:
                end = num
            else:
                if start == end:
                    ranges.append(str(start))
                else:
                    ranges.append(f"{start}-{end}")
                start = num
                end = num
        
        # Don't forget last range
        if start == end:
            ranges.append(str(start))
        else:
            ranges.append(f"{start}-{end}")
        
        return " / ".join(ranges)
    else:
        return " / ".join(str(n) for n in numbers)


def group_scans(scans: list) -> dict:
    """
    Group scans by Operator -> Station -> Part Number
    Returns nested dict with scan details
    """
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    
    for scan in scans:
        operator = scan.get('operator_name') or 'Unknown'
        station = scan.get('station_id') or 'Unknown'
        part = scan.get('part_id') or 'Unknown'
        serial = scan.get('serial_number') or ''
        
        grouped[operator][station][part].append(serial)
    
    return grouped


def generate_excel_report(grouped_data: dict, scans: list, report_date: datetime) -> str:
    """
    Generate Excel report matching Build_Report format.
    Returns path to temporary Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Build Summary {report_date.strftime('%m-%d-%Y')}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="072549", end_color="072549", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
    subtotal_fill = PatternFill(start_color="D9E8FB", end_color="D9E8FB", fill_type="solid")  # Light blue
    subtotal_font = Font(bold=True)
    
    # Collect data for tables
    table_rows = []
    part_header_totals = defaultdict(lambda: defaultdict(lambda: {'scans': 0, 'pieces': 0, 'serials': []}))
    
    for operator in grouped_data:
        for station in grouped_data[operator]:
            for part in grouped_data[operator][station]:
                serials = grouped_data[operator][station][part]
                scan_count = len(serials)
                pieces = scan_count * PIECES_PER_BOX
                serial_ranges = format_serial_ranges(serials)
                
                table_rows.append({
                    'operator': operator,
                    'station': station,
                    'part': part,
                    'scan_count': scan_count,
                    'pieces': pieces,
                    'serial_ranges': serial_ranges
                })
                
                # Track totals by part + serial header
                for serial in serials:
                    header = extract_serial_header(serial)
                    part_header_totals[part][header]['scans'] += 1
                    part_header_totals[part][header]['pieces'] += PIECES_PER_BOX
                    part_header_totals[part][header]['serials'].append(serial)
    
    # Sort by Part Number -> Operator -> Station
    table_rows.sort(key=lambda x: (x['part'], x['operator'], x['station']))
    
    # ===== MAIN SHEET: Build Summary by Part & Serial Header =====
    # Section Header
    ws.cell(row=1, column=1, value="Build Summary by Part & Serial Header").font = Font(bold=True, size=12)
    
    # Column headers
    gt_headers = ['Part Number', 'Total Scans (Boxes)', 'Total Pieces', 'Serial Header', 'Serial Numbers Logged', 'QB', 'SS']
    row = 2
    for col, header in enumerate(gt_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row += 1
    
    # Data rows with subtotals per part
    sorted_parts = sorted(part_header_totals.keys())
    row_counter = 0
    
    for part in sorted_parts:
        part_total_scans = 0
        part_total_pieces = 0
        headers_for_part = sorted(part_header_totals[part].keys())
        
        for header in headers_for_part:
            totals = part_header_totals[part][header]
            serial_ranges = format_serial_ranges(sorted(totals['serials']))
            
            row_fill = alt_row_fill if row_counter % 2 == 1 else None
            
            cell_part = ws.cell(row=row, column=1, value=part)
            cell_part.border = border
            cell_part.alignment = center_align
            if row_fill:
                cell_part.fill = row_fill
            
            cell_scans = ws.cell(row=row, column=2, value=totals['scans'])
            cell_scans.border = border
            cell_scans.alignment = center_align
            if row_fill:
                cell_scans.fill = row_fill
            
            cell_pieces = ws.cell(row=row, column=3, value=totals['pieces'])
            cell_pieces.border = border
            cell_pieces.alignment = center_align
            if row_fill:
                cell_pieces.fill = row_fill
            
            cell_header = ws.cell(row=row, column=4, value=header)
            cell_header.border = border
            cell_header.alignment = center_align
            if row_fill:
                cell_header.fill = row_fill
            
            cell_serial = ws.cell(row=row, column=5, value=serial_ranges)
            cell_serial.border = border
            cell_serial.alignment = wrap_align
            if row_fill:
                cell_serial.fill = row_fill
            
            cell_qb = ws.cell(row=row, column=6, value='')
            cell_qb.border = border
            cell_qb.alignment = center_align
            if row_fill:
                cell_qb.fill = row_fill
            
            cell_ss = ws.cell(row=row, column=7, value='')
            cell_ss.border = border
            cell_ss.alignment = center_align
            if row_fill:
                cell_ss.fill = row_fill
            
            part_total_scans += totals['scans']
            part_total_pieces += totals['pieces']
            
            row += 1
            row_counter += 1
        
        # Subtotal row (only if more than one serial header type)
        if len(headers_for_part) > 1:
            cell_sub_part = ws.cell(row=row, column=1, value=f"{part} SUBTOTAL")
            cell_sub_part.border = border
            cell_sub_part.alignment = center_align
            cell_sub_part.fill = subtotal_fill
            cell_sub_part.font = subtotal_font
            
            cell_sub_scans = ws.cell(row=row, column=2, value=part_total_scans)
            cell_sub_scans.border = border
            cell_sub_scans.alignment = center_align
            cell_sub_scans.fill = subtotal_fill
            cell_sub_scans.font = subtotal_font
            
            cell_sub_pieces = ws.cell(row=row, column=3, value=part_total_pieces)
            cell_sub_pieces.border = border
            cell_sub_pieces.alignment = center_align
            cell_sub_pieces.fill = subtotal_fill
            cell_sub_pieces.font = subtotal_font
            
            for col in range(4, 8):
                cell = ws.cell(row=row, column=col, value='')
                cell.border = border
                cell.fill = subtotal_fill
            
            row += 1
            row_counter = 0
    
    # ===== OPERATOR BREAKDOWN TABLE (on main sheet, below Build Summary) =====
    row += 2
    
    # Operator breakdown header
    ws.cell(row=row, column=1, value="Operator Breakdown").font = Font(bold=True, size=12)
    row += 1
    
    # Column headers for operator breakdown
    op_breakdown_headers = ['Operator', 'Station', 'Part Number', 'Total Pieces', 'Serial Numbers Logged', 'Context']
    for col, header in enumerate(op_breakdown_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row += 1
    
    op_row_idx = 0
    for data in table_rows:
        row_fill = alt_row_fill if op_row_idx % 2 == 1 else None
        
        cell = ws.cell(row=row, column=1, value=data['operator'])
        cell.border = border
        cell.alignment = center_align
        if row_fill:
            cell.fill = row_fill
        
        cell = ws.cell(row=row, column=2, value=data['station'])
        cell.border = border
        cell.alignment = center_align
        if row_fill:
            cell.fill = row_fill
        
        cell = ws.cell(row=row, column=3, value=data['part'])
        cell.border = border
        cell.alignment = center_align
        if row_fill:
            cell.fill = row_fill
        
        cell = ws.cell(row=row, column=4, value=data['pieces'])
        cell.border = border
        cell.alignment = center_align
        if row_fill:
            cell.fill = row_fill
        
        cell_serial = ws.cell(row=row, column=5, value=data['serial_ranges'])
        cell_serial.border = border
        cell_serial.alignment = wrap_align
        if row_fill:
            cell_serial.fill = row_fill
        
        cell = ws.cell(row=row, column=6, value='')
        cell.border = border
        cell.alignment = center_align
        if row_fill:
            cell.fill = row_fill
        
        row += 1
        op_row_idx += 1
    
    # ===== GRAND TOTAL BY OPERATOR (on main sheet, below Operator Breakdown) =====
    row += 2
    
    # Calculate operator totals
    operator_totals = {}
    for scan in scans:
        op = scan.get('operator_name', 'Unknown') or 'Unknown'
        if op not in operator_totals:
            operator_totals[op] = {'scans': 0, 'pieces': 0}
        operator_totals[op]['scans'] += 1
        operator_totals[op]['pieces'] += PIECES_PER_BOX
    
    # Operator totals header
    ws.cell(row=row, column=1, value="Grand Total by Operator").font = Font(bold=True, size=12)
    row += 1
    
    # Operator totals column headers
    op_tot_headers = ['Operator', 'Total Scans (Boxes)', 'Total Pieces']
    for col, header in enumerate(op_tot_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1
    
    # Operator totals data with alternating shading
    op_tot_row_idx = 0
    for op in sorted(operator_totals.keys()):
        row_fill = alt_row_fill if op_tot_row_idx % 2 == 1 else None
        
        cell = ws.cell(row=row, column=1, value=op)
        cell.border = border
        cell.alignment = center_align
        if row_fill:
            cell.fill = row_fill
        
        cell = ws.cell(row=row, column=2, value=operator_totals[op]['scans'])
        cell.border = border
        cell.alignment = center_align
        if row_fill:
            cell.fill = row_fill
        
        cell = ws.cell(row=row, column=3, value=operator_totals[op]['pieces'])
        cell.border = border
        cell.alignment = center_align
        if row_fill:
            cell.fill = row_fill
        
        row += 1
        op_tot_row_idx += 1

    # Auto-adjust column widths for main sheet with max cap
    max_widths = {'A': 24, 'B': 18, 'C': 14, 'D': 14, 'E': 60, 'F': 10, 'G': 8}
    for col_letter, max_width in max_widths.items():
        content_width = 10
        for cell in ws[col_letter]:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > content_width:
                    content_width = cell_len
        ws.column_dimensions[col_letter].width = min(content_width + 2, max_width)
    
    # Print settings
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    # ===== CREATE SUMMARY SHEET =====

    ws_summary = wb.create_sheet(title="Summary")
    
    # Title
    ws_summary.cell(row=1, column=1, value=f"Daily Build Assembly Report Summary - {report_date.strftime('%m-%d-%Y')}")
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    # Total scans
    total_scans = len(scans)
    ws_summary.cell(row=3, column=1, value="Total Scans:")
    ws_summary.cell(row=3, column=2, value=total_scans)
    ws_summary['A3'].font = Font(bold=True)
    
    ws_summary.cell(row=4, column=1, value="Total Pieces:")
    ws_summary.cell(row=4, column=2, value=total_scans * PIECES_PER_BOX)
    ws_summary['A4'].font = Font(bold=True)
    
    # By Station
    row = 6
    ws_summary.cell(row=row, column=1, value="By Station:")
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    station_counts = defaultdict(int)
    for scan in scans:
        station = scan.get('station_id') or 'Unknown'
        station_counts[station] += 1
    
    for station, count in sorted(station_counts.items(), key=lambda x: -x[1]):
        ws_summary.cell(row=row, column=1, value=station)
        ws_summary.cell(row=row, column=2, value=count)
        ws_summary.cell(row=row, column=3, value=f"({count * PIECES_PER_BOX} pcs)")
        row += 1
    
    # By Operator
    row += 1
    ws_summary.cell(row=row, column=1, value="By Operator:")
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    operator_counts = defaultdict(int)
    for scan in scans:
        operator = scan.get('operator_name') or 'Unknown'
        operator_counts[operator] += 1
    
    for operator, count in sorted(operator_counts.items(), key=lambda x: -x[1]):
        ws_summary.cell(row=row, column=1, value=operator)
        ws_summary.cell(row=row, column=2, value=count)
        ws_summary.cell(row=row, column=3, value=f"({count * PIECES_PER_BOX} pcs)")
        row += 1
    
    # Batch Comment (SO Context) Summary
    row += 2
    ws_summary.cell(row=row, column=1, value="Batch Comment ('SO' Context) Summary:")
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    # Count scans with and without SO context
    scans_with_so = 0
    scans_without_so = 0
    for scan in scans:
        comment = (scan.get('batch_comment') or '').upper()
        if 'SO' in comment:
            scans_with_so += 1
        else:
            scans_without_so += 1
    
    ws_summary.cell(row=row, column=1, value="Scans with 'SO' context:")
    ws_summary.cell(row=row, column=2, value=f"{scans_with_so} ({scans_with_so * PIECES_PER_BOX} pieces)")
    row += 1
    
    ws_summary.cell(row=row, column=1, value="Scans without 'SO' context:")
    ws_summary.cell(row=row, column=2, value=f"{scans_without_so} ({scans_without_so * PIECES_PER_BOX} pieces)")
    
    # Auto-adjust column widths for summary sheet
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    
    # ===== CREATE RAW DATA SHEET (for audit trail) =====
    ws_raw = wb.create_sheet(title="Raw Data")
    
    # Headers for raw data
    raw_headers = ['Timestamp (EST)', 'Serial Number', 'Part Number', 'Operator', 'Station', 'Raw Barcode', 'Comment', 'Notes']
    for col, header in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Populate raw data rows
    for row_num, scan in enumerate(scans, 2):
        # Convert UTC timestamp to EST for display
        ts = scan.get('created_at', '')
        if ts:
            try:
                dt = datetime.fromisoformat(ts.replace('+00:00', '').replace('Z', ''))
                ts_display = (dt - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')
            except:
                ts_display = ts
        else:
            ts_display = ''
        
        ws_raw.cell(row=row_num, column=1, value=ts_display).border = border
        ws_raw.cell(row=row_num, column=2, value=scan.get('serial_number', '')).border = border
        ws_raw.cell(row=row_num, column=3, value=scan.get('part_id', '')).border = border
        ws_raw.cell(row=row_num, column=4, value=scan.get('operator_name', '')).border = border
        ws_raw.cell(row=row_num, column=5, value=scan.get('station_id', '')).border = border
        ws_raw.cell(row=row_num, column=6, value=scan.get('raw_scan', '')).border = border
        ws_raw.cell(row=row_num, column=7, value=scan.get('batch_comment', '')).border = border
        ws_raw.cell(row=row_num, column=8, value=scan.get('dashboard_notes', '')).border = border
    
    # Auto-adjust column widths for raw data sheet
    raw_widths = [22, 18, 14, 12, 10, 45, 25, 25]
    for col, width in enumerate(raw_widths, 1):
        ws_raw.column_dimensions[get_column_letter(col)].width = width
    
    print(f"[OK] Added Raw Data sheet with {len(scans)} records")
    
    # ===== CREATE PART SUMMARY SHEET (Format requested: Part, Total Pcs, Ranges) =====
    ws_part_summary = wb.create_sheet(title="Part Summary")
    
    # Calculate global part totals first
    global_part_data = defaultdict(lambda: {'pieces': 0, 'serials': []})
    
    for scan in scans:
        part = scan.get('part_id') or 'Unknown'
        serial = scan.get('serial_number') or ''
        global_part_data[part]['pieces'] += PIECES_PER_BOX
        if serial:
            global_part_data[part]['serials'].append(serial)
    
    ps_row = 1
    # Check if we have data
    if not global_part_data:
        ws_part_summary.cell(row=1, column=1, value="No scans found.")
    else:
        for part in sorted(global_part_data.keys()):
            data = global_part_data[part]
            
            # Format:
            # <Part Number>
            # TOTAL PIECES: <Count>
            # SERIAL SEQUENCES: <Ranges>
            
            # Part Part Number
            cell_part = ws_part_summary.cell(row=ps_row, column=1, value=part)
            cell_part.font = Font(bold=True, size=12)
            ps_row += 1
            
            # Total Pieces
            ws_part_summary.cell(row=ps_row, column=1, value=f"TOTAL PIECES: {data['pieces']}")
            ps_row += 1
            
            # Serial Sequences
            ranges = format_serial_ranges(data['serials'])
            # Wrap text for long sequences
            cell_seq = ws_part_summary.cell(row=ps_row, column=1, value=f"SERIAL SEQUENCES: {ranges}")
            cell_seq.alignment = Alignment(wrap_text=True)
            ps_row += 1
            
            # Empty lines between parts
            ps_row += 2
            
    # Set column width
    ws_part_summary.column_dimensions['A'].width = 100
    
    print(f"[OK] Added Part Summary sheet")
    
    # ===== CREATE PER-OPERATOR SHEETS =====
    # Get unique operators from this day's scans
    operators = sorted(set(scan.get('operator_name', 'Unknown') or 'Unknown' for scan in scans))
    
    for operator in operators:
        # Filter scans for this operator
        operator_scans = [s for s in scans if (s.get('operator_name') or 'Unknown') == operator]
        
        if not operator_scans:
            continue
        
        # Create sheet (Excel sheet names max 31 chars, no special chars)
        sheet_name = f"{operator[:25]}"  # Truncate if too long
        # Remove any invalid characters
        sheet_name = ''.join(c for c in sheet_name if c not in '[]:*?/\\')
        if not sheet_name:
            sheet_name = "Unknown"
        
        ws_op = wb.create_sheet(title=sheet_name)
        
        # Headers
        op_headers = ['Timestamp (EST)', 'Serial Number', 'Part Number', 'Station', 'Raw Barcode', 'Comment']
        for col, header in enumerate(op_headers, 1):
            cell = ws_op.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Data rows
        for row_num, scan in enumerate(operator_scans, 2):
            # Convert timestamp to EST
            ts = scan.get('created_at', '')
            if ts:
                try:
                    dt = datetime.fromisoformat(ts.replace('+00:00', '').replace('Z', ''))
                    ts_display = (dt - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    ts_display = ts
            else:
                ts_display = ''
            
            ws_op.cell(row=row_num, column=1, value=ts_display).border = border
            ws_op.cell(row=row_num, column=2, value=scan.get('serial_number', '')).border = border
            ws_op.cell(row=row_num, column=3, value=scan.get('part_id', '')).border = border
            ws_op.cell(row=row_num, column=4, value=scan.get('station_id', '')).border = border
            ws_op.cell(row=row_num, column=5, value=scan.get('raw_scan', '')).border = border
            ws_op.cell(row=row_num, column=6, value=scan.get('batch_comment', '')).border = border
        
        # Column widths
        op_widths = [22, 18, 14, 10, 45, 25]
        for col, width in enumerate(op_widths, 1):
            ws_op.column_dimensions[get_column_letter(col)].width = width
        
        print(f"[OK] Added sheet for {operator} with {len(operator_scans)} scans")
    
    # Save to temp file
    temp_path = tempfile.mktemp(suffix='.xlsx')
    wb.save(temp_path)
    print(f"[OK] Generated Excel report: {temp_path}")
    
    return temp_path


def send_email(excel_path: str, report_date: datetime, scan_count: int):
    """Send email with Excel attachment"""
    recipients = [r.strip() for r in REPORT_RECIPIENTS.split(',')]
    
    if not SMTP_PASSWORD:
        print("[WARN] SMTP_PASSWORD not set, skipping email send")
        return
    
    msg = MIMEMultipart()
    msg['From'] = SMTP_EMAIL
    msg['To'] = ', '.join(recipients)
    msg['Subject'] = f"Daily Build Report - {report_date.strftime('%B %d, %Y')}"
    
    # Email body
    body = f"""
Good morning,

Please find attached the Daily Build Report for {report_date.strftime('%B %d, %Y')}.

Summary:
- Total Scans: {scan_count}
- Total Pieces: {scan_count * PIECES_PER_BOX:,}

This report was automatically generated from the Polytechnic Resources Serial Number Scan Log system.

Best regards,
Polytechnic Resources Automation
    """
    msg.attach(MIMEText(body, 'plain'))
    
    # Attach Excel file
    filename = f"Build_Report_{report_date.strftime('%m-%d-%Y')}.xlsx"
    with open(excel_path, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
    
    # Send
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_PASSWORD)
            server.sendmail(SMTP_EMAIL, recipients, msg.as_string())
        print(f"[OK] Email sent to: {', '.join(recipients)}")
    except Exception as e:
        print(f"[ERROR] Email failed: {e}")
        raise


def backup_to_google_sheets(scans: list, report_date: datetime):
    """
    Backup daily scans to a Google Sheet.
    Appends rows to the first worksheet with: Date, Timestamp, Serial, Part, Operator, Station, Comment
    """
    if not GSHEET_ENABLED:
        print("[WARN] Google Sheets backup not configured - skipping")
        return
    
    import base64
    import json
    
    try:
        # Decode service account credentials from base64
        creds_json = base64.b64decode(GSHEET_CREDENTIALS_JSON).decode('utf-8')
        creds_dict = json.loads(creds_json)
        
        # Authenticate with Google Sheets
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        gc = gspread.authorize(credentials)
        
        # Open spreadsheet
        spreadsheet = gc.open_by_key(GSHEET_SPREADSHEET_ID)
        worksheet = spreadsheet.sheet1  # First worksheet
        
        # Format rows for append
        date_str = report_date.strftime('%Y-%m-%d')
        rows_to_append = []
        
        for scan in scans:
            # Convert UTC timestamp to EST for display
            ts = scan.get('created_at', '')
            if ts:
                try:
                    dt = datetime.fromisoformat(ts.replace('+00:00', '').replace('Z', ''))
                    ts_display = (dt - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    ts_display = ts
            else:
                ts_display = ''
            
            rows_to_append.append([
                date_str,
                ts_display,
                scan.get('serial_number', ''),
                scan.get('part_id', ''),
                scan.get('operator_name', ''),
                scan.get('station_id', ''),
                scan.get('batch_comment', '')
            ])
        
        # Append all rows at once
        if rows_to_append:
            worksheet.append_rows(rows_to_append, value_input_option='RAW')
            print(f"[OK] Google Sheets backup: {len(rows_to_append)} rows appended")
        else:
            print("[WARN] No rows to backup to Google Sheets")
            
    except Exception as e:
        print(f"[ERROR] Google Sheets backup failed: {e}")
        # Don't raise - backup failure shouldn't stop the main report

def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Daily Build Report Generator')
    parser.add_argument('date', nargs='?', help='Report date (YYYY-MM-DD)')
    parser.add_argument('--test', action='store_true', help='Test mode (no email)')
    args = parser.parse_args()
    
    test_mode = args.test
    if args.date:
        try:
            target_date = datetime.strptime(args.date, '%Y-%m-%d')
        except ValueError:
            # Handle case where date might be mistakenly parsed if flags are mixed order in some shells
            # But argparse handles pos args well.
             print(f"[ERROR] Invalid date format: {args.date}. Use YYYY-MM-DD")
             sys.exit(1)
    else:
        target_date = datetime.now() - timedelta(days=1)
    
    print(f"[INFO] Generating Build Report for {target_date.strftime('%Y-%m-%d')}")
    print(f"   Test mode: {test_mode}")
    
    # Fetch data
    supabase = get_supabase_client()
    scans = fetch_scans_for_date(supabase, target_date)
    
    print(f"[INFO] Found {len(scans)} scans")
    
    if len(scans) == 0:
        print("[WARN] No scans found for this date. Skipping report.")
        return
    
    # Group data
    grouped = group_scans(scans)
    
    # Generate Excel (pass scans for summary sheet)
    excel_path = generate_excel_report(grouped, scans, target_date)
    
    # Send email (unless test mode)
    if not test_mode:
        send_email(excel_path, target_date, len(scans))
    else:
        print(f"[TEST] Test mode - Excel saved to: {excel_path}")
        print(f"   Would send to: {REPORT_RECIPIENTS}")
    
    # Backup to Google Sheets (always, including test mode)
    backup_to_google_sheets(scans, target_date)
    
    # Cleanup (unless test mode)
    if not test_mode:
        os.remove(excel_path)
    
    print("[OK] Report complete!")


if __name__ == '__main__':
    main()
